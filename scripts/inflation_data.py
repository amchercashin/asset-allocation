import json
import re
from io import BytesIO
from numbers import Real
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook


LANDING_URL = "https://rosstat.gov.ru/statistics/price"
ALLOWED_HOST = "rosstat.gov.ru"
WORKBOOK_RE = re.compile(r"ipc_mes_(\d{2})-(\d{4})\.xlsx$", re.IGNORECASE)
START_PERIOD = "2003-02"
MONTHS_RU = (
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
)


class InflationDataError(RuntimeError):
    pass


def discover_latest_workbook(html: bytes) -> str:
    text = html.decode("utf-8", errors="replace")
    candidates = []

    for href in re.findall(r'href=["\']([^"\']+)["\']', text, re.IGNORECASE):
        url = urljoin(LANDING_URL, href)
        parsed = urlparse(url)
        match = WORKBOOK_RE.search(parsed.path)
        if parsed.scheme != "https" or parsed.hostname != ALLOWED_HOST or not match:
            continue

        month, year = map(int, match.groups())
        if 1 <= month <= 12:
            candidates.append(((year, month), url))

    if not candidates:
        raise InflationDataError("Rosstat workbook link was not found")

    return max(candidates)[1]


def parse_workbook(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise InflationDataError(f"Could not open Rosstat workbook: {error}") from error

    try:
        if "01" not in workbook.sheetnames:
            raise InflationDataError("Rosstat workbook does not contain sheet 01")

        rows = list(workbook["01"].iter_rows(values_only=True))
        year_columns = _find_year_columns(rows)
        month_rows = _find_month_rows(rows)
        records = []
        series_ended = False

        for column, year in sorted(year_columns.items(), key=lambda item: item[1]):
            if year < 2003:
                continue
            for month_number, month_name in enumerate(MONTHS_RU, start=1):
                period = f"{year:04d}-{month_number:02d}"
                if period < START_PERIOD:
                    continue

                row = month_rows[month_name]
                value = row[column] if column < len(row) else None
                if value is None:
                    series_ended = True
                    continue
                if series_ended:
                    raise InflationDataError(
                        f"Rosstat workbook has data after a missing month at {period}"
                    )
                if isinstance(value, bool) or not isinstance(value, Real):
                    raise InflationDataError(
                        f"Rosstat value for {period} is not numeric: {value!r}"
                    )

                numeric = round(float(value), 2)
                if not 80 <= numeric <= 120:
                    raise InflationDataError(
                        f"Rosstat value for {period} is outside 80..120: {numeric}"
                    )
                records.append(
                    {
                        "period": period,
                        "index_to_previous_month": numeric,
                    }
                )

        if not records:
            raise InflationDataError("Rosstat workbook contains no monthly CPI data")
        return records
    finally:
        workbook.close()


def merge_months(
    existing: list[dict], official: list[dict]
) -> tuple[list[dict], bool]:
    validate_sequence(official)
    if not existing:
        return list(official), bool(official)

    validate_sequence(existing)
    if len(official) < len(existing):
        raise InflationDataError(
            "Official Rosstat series is older than local inflation data"
        )

    for local_row, official_row in zip(existing, official):
        if local_row["period"] != official_row["period"]:
            raise InflationDataError(
                "Official Rosstat historical period changed: "
                f"{local_row['period']} != {official_row['period']}"
            )
        local_value = round(float(local_row["index_to_previous_month"]), 2)
        official_value = round(float(official_row["index_to_previous_month"]), 2)
        if local_value != official_value:
            raise InflationDataError(
                "Official Rosstat historical value changed for "
                f"{local_row['period']}: {local_value} != {official_value}"
            )

    if len(official) == len(existing):
        return existing, False
    return list(official), True


def next_period(period: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{2})", period)
    if not match:
        raise InflationDataError(f"Invalid monthly period: {period!r}")
    year, month = map(int, match.groups())
    if not 1 <= month <= 12:
        raise InflationDataError(f"Invalid monthly period: {period!r}")
    if month == 12:
        return f"{year + 1:04d}-01"
    return f"{year:04d}-{month + 1:02d}"


def validate_sequence(rows: list[dict]) -> None:
    previous = None
    for row in rows:
        if not isinstance(row, dict):
            raise InflationDataError("Monthly CPI record must be an object")
        period = row.get("period")
        if not isinstance(period, str):
            raise InflationDataError("Monthly CPI period must be a string")
        next_period(period)

        value = row.get("index_to_previous_month")
        if isinstance(value, bool) or not isinstance(value, Real):
            raise InflationDataError(f"Monthly CPI value for {period} is not numeric")
        if not 80 <= float(value) <= 120:
            raise InflationDataError(
                f"Monthly CPI value for {period} is outside 80..120: {value}"
            )

        if previous is not None and period != next_period(previous):
            raise InflationDataError(
                f"Monthly CPI sequence is not consecutive: {previous}, {period}"
            )
        previous = period


def render_ipc_js(months: list[dict]) -> str:
    validate_sequence(months)
    records = [
        {
            "period": row["period"],
            "value": round(float(row["index_to_previous_month"]), 2),
        }
        for row in months
    ]
    monthly_json = json.dumps(records, ensure_ascii=False, separators=(",", ":"))

    return f'''// IPC (Consumer Price Index) data for Russia.
// Generated by scripts/update_inflation.py from the official Rosstat workbook.
// Do not edit manually; update data/inflation-monthly.json instead.

const monthlyCPI = {monthly_json};

let ipc = {{}};
ipc.x = [];
ipc.y = [];

let cumulative = 1.0;
let firstValue = null;

for (const record of monthlyCPI) {{
  cumulative *= record.value / 100;
  if (firstValue === null) {{
    firstValue = cumulative;
  }}

  const normalised = cumulative / firstValue;
  const [year, month] = record.period.split("-").map(Number);
  const day = new Date(year, month, 0).getDate();

  ipc.y.push(parseFloat(normalised.toFixed(10)));
  ipc.x.push(`${{record.period}}-${{String(day).padStart(2, "0")}}`);
}}

ipc.name = "ИПЦ";
ipc.type = "scatter";
ipc.line = {{}};
ipc.line.color = "grey";
ipc.line.dash = "solid";
'''


def _find_year_columns(rows: list[tuple]) -> dict[int, int]:
    best = {}
    for row in rows:
        candidate = {}
        for column, value in enumerate(row):
            if isinstance(value, bool) or not isinstance(value, Real):
                continue
            year = int(value)
            if value == year and 1900 <= year <= 2200:
                if year in candidate.values():
                    raise InflationDataError(f"Duplicate year {year} in Rosstat workbook")
                candidate[column] = year
        if len(candidate) > len(best):
            best = candidate

    if not best:
        raise InflationDataError("Could not find year headers in Rosstat workbook")
    return best


def _find_month_rows(rows: list[tuple]) -> dict[str, tuple]:
    found = {}
    for row in rows:
        for value in row:
            if not isinstance(value, str):
                continue
            month = value.strip().lower()
            if month in MONTHS_RU and month not in found:
                found[month] = row
                break

    missing = [month for month in MONTHS_RU if month not in found]
    if missing:
        raise InflationDataError(
            "Could not find month rows in Rosstat workbook: " + ", ".join(missing)
        )
    return found
